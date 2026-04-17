"""
Excel Export Service
Generates Excel exports of CBAM imports with formula breakdown and summary.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional, List
from uuid import UUID
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, extract, func
from pydantic import BaseModel

from app.models.uk_cbam import Import, UKCBAMProduct, Supplier


class ExportFilters(BaseModel):
    """Filters for Excel export."""
    year: Optional[int] = None
    sector: Optional[str] = None


def sanitize_excel_value(value):
    """
    Prevent Excel formula injection by prefixing dangerous characters.
    
    Args:
        value: Cell value
    
    Returns:
        Sanitized value with single quote prefix if needed
    """
    if value and isinstance(value, str) and len(value) > 0:
        if value[0] in ('=', '+', '-', '@'):
            return "'" + value
    return value


def format_currency(value: Optional[Decimal]) -> str:
    """Format decimal as currency string."""
    if value is None:
        return "£0.00"
    return f"£{value:,.2f}"


def format_number(value: Optional[Decimal], decimals: int = 2) -> str:
    """Format decimal as number string."""
    if value is None:
        return "0.00"
    format_str = f"{{:,.{decimals}f}}"
    return format_str.format(value)


async def generate_excel_export(
    org_id: UUID,
    filters: ExportFilters,
    db: AsyncSession
) -> bytes:
    """
    Generate Excel export of imports with formula breakdown and summary.
    
    Args:
        org_id: Organisation UUID
        filters: Export filters (year, sector)
        db: Database session
    
    Returns:
        Excel file as bytes
    
    Algorithm:
        1. Query imports with filters
        2. Create workbook with 3 sheets
        3. Generate "Imports Summary" sheet
        4. Generate "Formula Breakdown" sheet
        5. Generate "Summary" sheet with aggregations
        6. Apply formatting
        7. Return as bytes
    
    Preconditions:
        - org_id exists in organisations table
        - db session is active
        - At least one import exists for org_id (after filters applied)
    
    Postconditions:
        - Returns valid .xlsx file as bytes
        - File contains exactly 3 sheets
        - All imports matching filters are included
        - Numeric values are properly formatted
        - File size < 10MB
    """
    # Build query
    query = select(Import).where(
        and_(
            Import.org_id == org_id,
            Import.is_deleted == False
        )
    )
    
    # Apply filters
    if filters.year:
        query = query.where(extract('year', Import.import_date) == filters.year)
    
    if filters.sector:
        query = query.join(UKCBAMProduct).where(UKCBAMProduct.sector == filters.sector)
    
    # Order by date descending
    query = query.order_by(Import.import_date.desc())
    
    # Execute query
    result = await db.execute(query)
    imports = result.scalars().all()
    
    if not imports:
        raise ValueError("No imports found matching filters")
    
    # Load related data
    import_ids = [imp.id for imp in imports]
    
    # Load products
    product_ids = [imp.product_id for imp in imports]
    products_result = await db.execute(
        select(UKCBAMProduct).where(UKCBAMProduct.id.in_(product_ids))
    )
    products = {p.id: p for p in products_result.scalars().all()}
    
    # Load suppliers
    supplier_ids = [imp.supplier_id for imp in imports if imp.supplier_id]
    suppliers = {}
    if supplier_ids:
        suppliers_result = await db.execute(
            select(Supplier).where(Supplier.id.in_(supplier_ids))
        )
        suppliers = {s.id: s for s in suppliers_result.scalars().all()}
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Imports Summary
    ws_summary = wb.create_sheet("Imports Summary")
    _create_imports_summary_sheet(ws_summary, imports, products, suppliers)
    
    # Sheet 2: Formula Breakdown
    ws_formula = wb.create_sheet("Formula Breakdown")
    _create_formula_breakdown_sheet(ws_formula, imports, products)
    
    # Sheet 3: Summary Statistics
    ws_stats = wb.create_sheet("Summary")
    _create_summary_sheet(ws_stats, imports, products)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def _create_imports_summary_sheet(
    ws,
    imports: List[Import],
    products: dict,
    suppliers: dict
):
    """Create the Imports Summary sheet."""
    # Headers
    headers = [
        "Import Date", "Product", "CN Code", "Sector", "Quantity (t)",
        "Country", "Supplier", "Intensity Used (tCO₂e/t)",
        "Embedded Emissions (tCO₂e)", "UK ETS Rate (£/tCO₂e)",
        "CBAM Liability (£)", "Default Liability (£)",
        "Savings (£)", "Data Source", "Import Type"
    ]
    
    # Header row styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_num, imp in enumerate(imports, 2):
        product = products.get(imp.product_id)
        supplier = suppliers.get(imp.supplier_id) if imp.supplier_id else None
        
        intensity_used = imp.emissions_intensity_actual if imp.emissions_intensity_actual else imp.emissions_intensity_default
        
        row_data = [
            imp.import_date.strftime('%Y-%m-%d') if imp.import_date else '',
            sanitize_excel_value(product.description if product else ''),
            sanitize_excel_value(product.commodity_code if product else ''),
            sanitize_excel_value(product.sector if product else ''),
            float(imp.quantity_tonnes) if imp.quantity_tonnes else 0,
            sanitize_excel_value(imp.country_of_origin),
            sanitize_excel_value(supplier.name if supplier else '—'),
            float(intensity_used) if intensity_used else 0,
            float(imp.embedded_emissions_tco2e) if imp.embedded_emissions_tco2e else 0,
            float(imp.uk_ets_rate_used) if imp.uk_ets_rate_used else 0,
            float(imp.cbam_liability_gbp) if imp.cbam_liability_gbp else 0,
            float(imp.cbam_liability_default_gbp) if imp.cbam_liability_default_gbp else 0,
            float(imp.potential_saving_gbp) if imp.potential_saving_gbp else 0,
            sanitize_excel_value(imp.data_source),
            sanitize_excel_value(imp.import_type)
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Format numeric columns
            if col_num in [5, 8, 9, 10]:  # Quantity, emissions, rate
                cell.number_format = '#,##0.00'
            elif col_num in [11, 12, 13]:  # Currency columns
                cell.number_format = '£#,##0.00'
    
    # Auto-size columns
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15


def _create_formula_breakdown_sheet(
    ws,
    imports: List[Import],
    products: dict
):
    """Create the Formula Breakdown sheet."""
    # Headers
    headers = ["Import ID", "Product", "Formula", "Calculation Steps"]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_num, imp in enumerate(imports, 2):
        product = products.get(imp.product_id)
        
        # Build formula string
        intensity = imp.emissions_intensity_actual if imp.emissions_intensity_actual else imp.emissions_intensity_default
        formula = f"Liability = (Quantity × Intensity × UK ETS Rate) - Deduction"
        
        # Build calculation steps
        steps = []
        steps.append(f"1. Embedded Emissions = {imp.quantity_tonnes} t × {intensity} tCO₂e/t = {imp.embedded_emissions_tco2e} tCO₂e")
        gross = (imp.cbam_liability_gbp or 0) + (imp.carbon_price_deduction_gbp or 0)
        steps.append(f"2. Gross Liability = {imp.embedded_emissions_tco2e} tCO₂e × £{imp.uk_ets_rate_used}/tCO₂e = £{gross}")
        if imp.carbon_price_deduction_gbp and imp.carbon_price_deduction_gbp > 0:
            steps.append(f"3. Deduction = £{imp.carbon_price_deduction_gbp}")
            steps.append(f"4. Net Liability = £{imp.cbam_liability_gbp}")
        else:
            steps.append(f"3. Net Liability = £{imp.cbam_liability_gbp} (no deduction)")
        
        if imp.potential_saving_gbp and imp.potential_saving_gbp > 0:
            steps.append(f"5. Potential Saving = £{imp.cbam_liability_default_gbp} - £{imp.cbam_liability_gbp} = £{imp.potential_saving_gbp}")
        
        calculation_text = "\n".join(steps)
        
        ws.cell(row=row_num, column=1, value=str(imp.id))
        ws.cell(row=row_num, column=2, value=sanitize_excel_value(product.description if product else ''))
        ws.cell(row=row_num, column=3, value=formula)
        ws.cell(row=row_num, column=4, value=calculation_text)
        
        # Set row height for multi-line text
        ws.row_dimensions[row_num].height = 80
        ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    
    # Column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60


def _create_summary_sheet(
    ws,
    imports: List[Import],
    products: dict
):
    """Create the Summary statistics sheet."""
    # Calculate totals
    total_imports = len(imports)
    total_liability = sum(imp.cbam_liability_gbp or 0 for imp in imports)
    total_savings = sum(imp.potential_saving_gbp or 0 for imp in imports)
    total_emissions = sum(imp.embedded_emissions_tco2e or 0 for imp in imports)
    avg_liability = total_liability / total_imports if total_imports > 0 else 0
    
    # Header styling
    header_font = Font(bold=True, size=14)
    metric_font = Font(bold=True)
    
    # Overall metrics
    ws.cell(row=1, column=1, value="Overall Metrics").font = header_font
    
    metrics = [
        ("Total Imports", total_imports),
        ("Total CBAM Liability", f"£{total_liability:,.2f}"),
        ("Total Potential Savings", f"£{total_savings:,.2f}"),
        ("Total Embedded Emissions", f"{total_emissions:,.2f} tCO₂e"),
        ("Average Liability per Import", f"£{avg_liability:,.2f}"),
    ]
    
    for idx, (metric, value) in enumerate(metrics, 2):
        ws.cell(row=idx, column=1, value=metric).font = metric_font
        ws.cell(row=idx, column=2, value=value)
    
    # Breakdown by sector
    row_offset = len(metrics) + 3
    ws.cell(row=row_offset, column=1, value="Breakdown by Sector").font = header_font
    
    # Group by sector
    sector_data = {}
    for imp in imports:
        product = products.get(imp.product_id)
        if product:
            sector = product.sector
            if sector not in sector_data:
                sector_data[sector] = {
                    'count': 0,
                    'liability': Decimal('0'),
                    'emissions': Decimal('0')
                }
            sector_data[sector]['count'] += 1
            sector_data[sector]['liability'] += imp.cbam_liability_gbp or 0
            sector_data[sector]['emissions'] += imp.embedded_emissions_tco2e or 0
    
    # Sector headers
    row_offset += 1
    ws.cell(row=row_offset, column=1, value="Sector").font = metric_font
    ws.cell(row=row_offset, column=2, value="Imports").font = metric_font
    ws.cell(row=row_offset, column=3, value="Total Liability").font = metric_font
    ws.cell(row=row_offset, column=4, value="Total Emissions").font = metric_font
    
    # Sector data
    for sector, data in sorted(sector_data.items()):
        row_offset += 1
        ws.cell(row=row_offset, column=1, value=sanitize_excel_value(sector))
        ws.cell(row=row_offset, column=2, value=data['count'])
        ws.cell(row=row_offset, column=3, value=f"£{data['liability']:,.2f}")
        ws.cell(row=row_offset, column=4, value=f"{data['emissions']:,.2f} tCO₂e")
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
